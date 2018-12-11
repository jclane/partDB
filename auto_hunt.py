from openpyxl import load_workbook, Workbook
import datetime as dt
from shutil import copyfile
from csv import reader as csvreader
from csv import writer as csvwriter
from os.path import join as pathjoin

from backend import search_part, convert_to_dict


def get_list(part_type):
    """
    Returns a list of part numbers based on part_type.

    :param part_type: String indicating 'HDD', 'MEM', or 'CPU'
    :return: List of part numbers from csv file
    """
    file = pathjoin("parts_in_sp", "all" + part_type + ".csv")
    data = []
    with open(file, "r") as csvfile:
        reader = csvreader(csvfile)
        for row in reader:
            data.append(row[0])
    return data


def copy_file(original):
    """
    Copies the openPO report to the local HDD and returns that
    copy.

    :param original: Path of the file to be copied
    :return: Path of the newly created local copy
    """
    local_copy = r"[REDACTED]" + \
                 str(dt.date.today()) + ".xlsx"

    copyfile(original, local_copy)
    return local_copy


def get_type(part_num):
    """
    Checks if part_num is in all_hdds, all_mem, or all_cpus and
    returns a sring indicating which is true.

    :param part_num: Part number
    :return: "HDD", "MEM", or "CPU"
    """
    all_hdds = get_list("HDD")
    all_mem = get_list("MEM")
    all_cpus = get_list("CPU")

    if part_num in all_hdds:
        return "HDD"
    elif part_num in all_mem:
        return "MEM"
    elif part_num in all_cpus:
        return "CPU"
    else:
        return None


def purge_subbed(part_nums):
    """
    Loops through part_nums and checks if the part number is
    in the database and if not, adds it to the clean list. Also
    calls convert_to_dict to check if part number has a sub.
    If not, the part is added to the clean_list.

    :param part_nums: List of part numbers to be checked
    :return: List of parts that are not in the databse or have no
        sub relation setup
    """
    clean_list = []
    for part_num in part_nums:
        if (search_part(part_num[1].lower(), part_num[0]) is None and
                part_num not in clean_list):
            part_num.append("+DB")
            clean_list.append(part_num)
        elif (convert_to_dict(part_num[1].lower(), part_num[0])
              ["subbed"] == "FALSE"):
            part_num.append("+SUB")
            clean_list.append(part_num)

    return clean_list


def save_to_file(part_nums):
    """
    Creates a workbook object and builds the 'Needs Sub' and
    'Non-Warr Orders' sheets then saves the workbook as an
    Excel file.

    :param part_nums: List of part numbers to be saved
    :return: 'Done' or error message if file is open
    """

    def build_sheets():
        """
        Populates the sheets with values based on the data provided
        in the part_nums list.
        """
        parts_seen = []
        add_to_sheet2 = []

        # Populating cell data for 'Needs Sub' sheet
        headings = ["Part Num", "Type"]
        for heading in enumerate(headings):
            worksheet1.cell(row=1, column=heading[0] + 1,
                            value=heading[1])

        row_num = 2
        for row in part_nums:
            if row[2] != "MFG Warranty":
                add_to_sheet2.append(row)
            if row[0] not in parts_seen:
                for part_data in enumerate(row[:2]):
                    worksheet1.cell(row=row_num, column=part_data[0] + 1,
                                    value=part_data[1])
                parts_seen.append(row[0])
                row_num += 1
            else:
                continue
        # Populating cell data for 'Non-Warr Orders' sheet
        if len(add_to_sheet2):
            headings = ["Part Num", "Type", "Warr", "SO"]
            for heading in enumerate(headings):
                worksheet2.cell(row=1, column=heading[0] + 1,
                                value=heading[1])
            for row in enumerate(add_to_sheet2):
                for element in enumerate(row[1][:-1]):
                    worksheet2.cell(row=row[0] + 2, column=element[0] + 1,
                                    value=element[1])

    path = pathjoin(
        r"[REDACTED]\hunts",
        str(dt.date.today()) + " hunt.xlsx"
        )
    workbook = Workbook()
    worksheet1 = workbook.active
    worksheet1.title = "Needs Sub"
    worksheet2 = workbook.create_sheet("Non-Warr Orders")
    build_sheets()

    workbook.save(path)
    return "Done"
